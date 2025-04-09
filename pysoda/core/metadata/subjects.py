from .constants import METADATA_UPLOAD_PS_PATH, TEMPLATE_PATH, SDS_FILE_SUBJECTS,SCHEMA_NAME_SUBJECTS
from .excel_utils import rename_headers, excel_columns
from openpyxl.styles import PatternFill
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
import numpy as np
from ...utils import validate_schema
from openpyxl.styles import Font
from .helpers import transposeMatrix, getMetadataCustomFields, sortedSubjectsTableData, upload_metadata_file


def create_excel(soda, upload_boolean, local_destination):
    source = join(TEMPLATE_PATH, SDS_FILE_SUBJECTS)


    datastructure = soda["dataset-metadata"]["subjects"]

    validate_schema(datastructure, SCHEMA_NAME_SUBJECTS)

    destination = join(METADATA_UPLOAD_PS_PATH, SDS_FILE_SUBJECTS) if upload_boolean else local_destination
    shutil.copyfile(source, destination)
    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]

    transposeDatastructure = transposeMatrix(datastructure)

    mandatoryFields = transposeDatastructure[:11]
    optionalFields = transposeDatastructure[11:]
    refinedOptionalFields = getMetadataCustomFields(optionalFields)

    templateHeaderList = subjectsTemplateHeaderList
    sortMatrix = sortedSubjectsTableData(mandatoryFields, templateHeaderList)

    if refinedOptionalFields:
        refinedDatastructure = transposeMatrix(
            np.concatenate((sortMatrix, refinedOptionalFields))
        )
    else:
        refinedDatastructure = transposeMatrix(sortMatrix)
    
    # delete all optional columns first (from the template)
    ws1.delete_cols(12, 18)

    # 1. see if the length of datastructure[0] == length of datastructure. If yes, go ahead. If no, add new columns from headers[n-1] onward.
    headers_no = len(refinedDatastructure[0])
    orangeFill = PatternFill(
        start_color="FFD965", end_color="FFD965", fill_type="solid"
    )

    for column, header in zip(
        excel_columns(start_index=11), refinedDatastructure[0][11:headers_no]
    ):
        cell = column + str(1)
        ws1[cell] = header
        ws1[cell].fill = orangeFill
        ws1[cell].font = Font(bold=True, size=12, name="Calibri")

    # 2. populate matrices
    for i, item in enumerate(refinedDatastructure):
        if i == 0:
            continue
        for column, j in zip(excel_columns(start_index=0), range(len(item))):
            # import pdb; pdb.set_trace()
            cell = column + str(i + 1)
            ws1[cell] = refinedDatastructure[i][j] or ""
            ws1[cell].font = Font(bold=False, size=11, name="Arial")

    wb.save(destination)

    size = getsize(destination)

    ## if generating directly on Pennsieve, then call upload function and then delete the destination path
    if upload_boolean:
        upload_metadata_file(SDS_FILE_SUBJECTS, soda, destination, True)

    return size



subjectsTemplateHeaderList = [
    "subject id",
    "pool id",
    "subject experimental group",
    "age",
    "sex",
    "species",
    "strain",
    "rrid for strain",
    "age category",
    "also in dataset",
    "member of",
    "metadata only",
    "laboratory internal id",
    "date of birth",
    "age range (min)",
    "age range (max)",
    "body mass",
    "genotype",
    "phenotype",
    "handedness",
    "reference atlas",
    "experimental log file path",
    "experiment date",
    "disease or disorder",
    "intervention",
    "disease model",
    "protocol title",
    "protocol url or doi",
]




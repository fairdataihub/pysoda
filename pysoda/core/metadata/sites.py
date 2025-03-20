from .constants import METADATA_UPLOAD_BF_PATH, TEMPLATE_PATH, SDS_FILE_SITES, SCHEMA_NAME_SITES
from .excel_utils import rename_headers, excel_columns
from openpyxl.styles import PatternFill
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
from ...utils import validate_schema

def create_excel(soda, upload_boolean, filepath):
    source = join(TEMPLATE_PATH, SDS_FILE_SITES)

    if upload_boolean:
        destination = join(METADATA_UPLOAD_BF_PATH, SDS_FILE_SITES)
    else:
        destination = filepath

    shutil.copyfile(source, destination)

    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]

    sites = soda["dataset_metadata"]["sites"]

    validate_schema(sites, SCHEMA_NAME_SITES)


    # get the ascii column headers
    row = 2
    ascii_headers = excel_columns(start_index=0)
    for performance in sites: 
        ws1[ascii_headers[0] + str(row)] = performance.get("site_id", "")
        ws1[ascii_headers[1] + str(row)] = performance.get("specimen_id", "")
        ws1[ascii_headers[2] + str(row)] = performance.get("site_type", "")
        ws1[ascii_headers[3] + str(row)] = performance.get("laboratory_internal_id", "")
        ws1[ascii_headers[4] + str(row)] = performance.get("coordinate_system", "")
        ws1[ascii_headers[5] + str(row)] = performance.get("coordinate_system_position", "")
        ws1[ascii_headers[6] + str(row)] = performance.get("more...", "")
        row += 1

    wb.save(destination)




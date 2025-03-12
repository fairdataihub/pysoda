from os.path import join, getsize, abspath, dirname
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from string import ascii_uppercase
import itertools
import shutil
import pkg_resources
import json
from jsonschema import validate

TEMPLATE_PATH = join(dirname(abspath(__file__)), '..', 'metadata_templates')
METADATA_UPLOAD_BF_PATH = join(dirname(abspath(__file__)), 'metadata_upload_bf')




def load_schema(schema_name):
    schema_path  = pkg_resources.resource_filename(__name__, f'../../../schema/{schema_name}')
    with open(schema_path, 'r') as schema_file:
        schema = json.load(schema_file)
    return schema

def validate_submission_metadata(submission_metadata):
    """
    Validate submission metadata against the submission schema.

    Args:
        submission_metadata (dict): The submission metadata to validate.

    Raises:
        ValidationError: If the submission metadata is invalid.
    """
    schema = load_schema('submission_schema.json')
    validate(instance=submission_metadata, schema=schema)

   
### Create submission file
def create_excel(soda, upload_boolean, destination_path):
    """
    Create an Excel file for submission metadata.

    Args:
        soda (dict): The soda object containing dataset metadata.
        upload_boolean (bool): Whether to upload the file to Pennsieve.
        destination_path (str): The path to save the Excel file.

    Returns:
        dict: A dictionary containing the size of the metadata file.
    """

    validate_submission_metadata(soda["dataset_metadata"]["submission_metadata"])

    font_submission = Font(name="Calibri", size=14, bold=False)

    source = join(TEMPLATE_PATH, "submission.xlsx")

    destination = join(METADATA_UPLOAD_BF_PATH, "submission.xlsx") if upload_boolean else destination_path

    try:
        shutil.copyfile(source, destination)
    except FileNotFoundError as e:
        raise e
    
    #TODO: Do not use an array for the non-array values; zipping for the sake of the ascii value is not necessary until milestone_achieved
    submission_metadata_list = [
        soda["dataset_metadata"]["submission_metadata"]
    ]

    # write to excel file
    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]
    start_index = 2
    for column, submission_data in zip(excel_columns(start_index), submission_metadata_list):
        ws1[column + "2"] = submission_data["consortium_data_standard"]
        ws1[column + "3"] = submission_data["funding_consortium"]
        ws1[column + "4"] = submission_data["award_number"]
        for col, milestone in zip(excel_columns(start_index), submission_data["milestone_achieved"]):
            ws1[col + str(5)] = milestone
        ws1[column + "6"] = submission_data["milestone_completion_date"]
        ws1[column + "2"].font = font_submission
        ws1[column + "3"].font = font_submission
        ws1[column + "4"].font = font_submission
        ws1[column + "5"].font = font_submission
        ws1[column + "6"].font = font_submission

    # TODO: should milestone completion date also be an array?
    rename_headers(ws1, len(submission_metadata_list[0]["milestone_achieved"]), 2)

    wb.save(destination)

    wb.close()

    # calculate the size of the metadata file
    size = getsize(destination)

    

    ## if generating directly on Pennsieve, then call upload function and then delete the destination path
    if upload_boolean:
        print("Implement later")
        # upload_metadata_file("submission.xlsx", bfaccount, bfdataset, destination, True)
    return {"size": size}

def excel_columns(start_index=0):
    """
    NOTE: does not support more than 699 contributors/links
    """
    single_letter = list(ascii_uppercase[start_index:])
    two_letter = [a + b for a, b in itertools.product(ascii_uppercase, ascii_uppercase)]
    return single_letter + two_letter


def rename_headers(workbook, max_len, start_index):
  """
  Rename header columns if values exceed 3. Change Additional Values to Value 4, 5,...
  Adds styling to the column headers as well.
  """

  columns_list = excel_columns(start_index=start_index)
  if max_len >= start_index:
      workbook[columns_list[0] + "1"] = "Value"
      for i, column in zip(range(2, max_len + 1), columns_list[1:]):

          workbook[column + "1"] = f"Value {str(i)}"
          cell = workbook[column + "1"]

          blueFill = PatternFill(
              start_color="9CC2E5", end_color="9CC2E5", fill_type="solid"
          )

          font = Font(bold=True)
          cell.fill = blueFill
          cell.font = font

  else:
      delete_range = len(columns_list) - max_len
      workbook.delete_cols(4 + max_len, delete_range)








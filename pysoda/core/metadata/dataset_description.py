from .constants import METADATA_UPLOAD_PS_PATH, TEMPLATE_PATH, SDS_FILE_DATASET_DESCRIPTION, SCHEMA_NAME_DATASET_DESCRIPTION
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
from .excel_utils import rename_headers, excel_columns
import itertools
from openpyxl.styles import PatternFill
from ...utils import validate_schema
from .helpers import upload_metadata_file, get_template_path



def create_excel(
    soda,
    upload_boolean,
    local_destination,
):
    source = get_template_path(SDS_FILE_DATASET_DESCRIPTION)
    destination = join(METADATA_UPLOAD_PS_PATH, SDS_FILE_DATASET_DESCRIPTION) if upload_boolean else local_destination
    shutil.copyfile(source, destination)

    validate_schema(soda["dataset_metadata"]["dataset_description"], SCHEMA_NAME_DATASET_DESCRIPTION)

    # write to excel file
    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]

    ws1["D22"] = ""
    ws1["E22"] = ""
    ws1["D24"] = ""
    ws1["E24"] = ""
    ws1["D25"] = ""
    ws1["E25"] = ""

    # Populate the Metadata version (Required)
    ws1["D2"] = soda["dataset_metadata"]["dataset_description"]["metadata_version"]

    # Populate the Dataset Type (default to empty string if not present)
    ws1["D3"] = (
    soda.get("dataset_metadata", {})
        .get("dataset_description", {})
        .get("dataset_type", "")
    )

    populate_standards_info(ws1, soda)

    keyword_array = populate_basic_info(ws1, soda)

    populate_study_info(ws1, soda)
    populate_contributor_info(ws1, soda)
    populate_related_resource_information(ws1, soda)
    populate_funding_info(ws1, soda)
    populate_participant_information(ws1, soda)
    data_dictionary_information(ws1, soda)

    wb.save(destination)

    size = getsize(destination)

    ## if generating directly on Pennsieve, then call upload function and then delete the destination path
    if upload_boolean:
        upload_metadata_file(
            "dataset_description.xlsx", soda, destination, True
        )

    return {"size": size}




def populate_study_info(workbook, soda):
    study_info = soda["dataset_metadata"]["dataset_description"]["study_information"]
    workbook["D20"] = study_info.get("study_purpose", "")
    workbook["D21"] = study_info.get("study_data_collection", "")
    workbook["D22"] = study_info.get("study_primary_conclusion", "")

    # Arrays
    organ_system = study_info.get("study_organ_system", [])
    approach = study_info.get("study_approach", [])
    technique = study_info.get("study_technique", [])

    for i, column in zip(range(len(organ_system)), excel_columns(start_index=3)):
        workbook[column + "23"] = organ_system[i]
    for i, column in zip(range(len(approach)), excel_columns(start_index=3)):
        workbook[column + "24"] = approach[i]
    for i, column in zip(range(len(technique)), excel_columns(start_index=3)):
        workbook[column + "25"] = technique[i]

    workbook["D26"] = study_info.get("study_collection_title", "")

    # Return the max length of the arrays, or 1 if all are empty
    return max(1, len(organ_system), len(approach), len(technique))

def populate_standards_info(workbook, soda):
    standards_info = soda["dataset_metadata"]["dataset_description"]["standards_information"]
    workbook["D5"] = standards_info["data_standard"]
    workbook["D6"] = standards_info["data_standard_version"]


def populate_basic_info(workbook, soda):
    basic_info = soda["dataset_metadata"]["dataset_description"]["basic_information"]
    workbook["D8"] = basic_info.get("title", "")
    workbook["D9"] = basic_info.get("subtitle", "")
    workbook["D10"] = basic_info.get("description", "")

    # Write keywords array across columns in row 11 (D11, E11, F11, ...)
    keywords = basic_info.get("keywords", [])
    for col, keyword in zip(excel_columns(start_index=3), keywords):
        workbook[f"{col}11"] = keyword

    workbook["D12"] = basic_info.get("funding", "")
    workbook["D13"] = basic_info.get("acknowledgments", "")
    workbook["D14"] = basic_info.get("license", "")

    # Return the length of the keywords array, or 1 if empty
    return max(1, len(keywords))


def populate_funding_info(workbook, soda):
    funding_info = soda["dataset_metadata"]["dataset_description"]["funding_information"]
    workbook["D16"] = funding_info["funding_consortium"]
    workbook["D17"] = funding_info["funding_agency"]
    workbook["D18"] = funding_info["award_number"]



def populate_contributor_info(workbook, soda):
    contributor_info = soda["dataset_metadata"]["dataset_description"].get("contributor_information", [])
    for contributor, column in zip(contributor_info, excel_columns(start_index=3)):
        workbook[column + "28"] = contributor.get("contributor_name", "")
        workbook[column + "29"] = contributor.get("contributor_orcid_id", "")
        workbook[column + "30"] = contributor.get("contributor_affiliation", "")
        workbook[column + "31"] = contributor.get("contributor_role", "")
    # Return the length of the contributor array, or 1 if empty
    return max(1, len(contributor_info))


def populate_related_resource_information(workbook, soda):
    related_resource_information = soda["dataset_metadata"]["dataset_description"].get("related_resource_information", [])
    for info, column in zip(related_resource_information, excel_columns(start_index=3)):
        workbook[column + "33"] = info.get("identifier_description", "")
        workbook[column + "34"] = info.get("relation_type", "")
        workbook[column + "35"] = info.get("identifier", "")
        workbook[column + "36"] = info.get("identifier_type", "")
    # Return the length of the related resource array, or 1 if empty
    return max(1, len(related_resource_information))



def populate_participant_information(workbook, soda):
    participant_info = soda["dataset_metadata"]["dataset_description"]["participant_information"]
    workbook["D38"] = participant_info.get("number_of_subjects", 0)
    workbook["D39"] = participant_info.get("number_of_samples", 0)
    workbook["D40"] = participant_info.get("number_of_sites", 0)
    workbook["D41"] = participant_info.get("number_of_performances", 0)


def data_dictionary_information(workbook, soda):
    """
    This function is a placeholder for future implementation.
    It currently does not populate any data in the workbook.
    """
    # Placeholder for future implementation
    data_dictionary_info = soda["dataset_metadata"]["dataset_description"].get("data_dictionary_information", {})

    workbook["D43"] = data_dictionary_info.get("data_dictionary_path", "")
    workbook["D44"] = data_dictionary_info.get("data_dictionary_type", "")
    workbook["D45"] = data_dictionary_info.get("data_dictionary_description", "")

def grayout_subheaders(workbook, max_len, start_index):
    """
    Gray out sub-header rows for values exceeding 3 (SDS2.0).
    """
    headers_list = ["4", "10", "18", "23", "28"]
    columns_list = excel_columns(start_index=start_index)

    for (i, column), no in itertools.product(zip(range(2, max_len + 1), columns_list[1:]), headers_list):
        cell = workbook[column + no]
        fillColor("B2B2B2", cell)





def fillColor(color, cell):
    colorFill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    cell.fill = colorFill

from .constants import METADATA_UPLOAD_PS_PATH, TEMPLATE_PATH, SDS_FILE_MANIFEST, SCHEMA_NAME_MANIFEST
from .excel_utils import rename_headers, excel_columns
from openpyxl.styles import Font, PatternFill
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
from ...utils import validate_schema
from .helpers import upload_metadata_file
import os
import sys


from json import load as json_load

def get_template_path(filename):
    """Get the path to a template file within the metadata_templates package."""
    
    # Method 1: Try PyInstaller bundle first (onefolder creates _MEIPASS)
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller onefolder extracts to _MEIPASS/
        possible_paths = [
            os.path.join(sys._MEIPASS, "pysoda", "core", "metadata_templates", filename),
            os.path.join(sys._MEIPASS, "metadata_templates", filename),
            os.path.join(sys._MEIPASS, filename)
        ]
        for path in possible_paths:
            if os.path.exists(path):
                return path
    
    # Method 2: Try to import the metadata_templates module (works if PyPI package is properly installed)
    try:
        from .. import metadata_templates
        templates_dir = os.path.dirname(metadata_templates.__file__)
        template_path = os.path.join(templates_dir, filename)
        if os.path.exists(template_path):
            return template_path
    except (ImportError, ModuleNotFoundError, AttributeError):
        pass
    
    # Method 3: Search in the Flask app's directory structure
    current_file = os.path.abspath(__file__)
    current_dir = os.path.dirname(current_file)
    
    # Walk up the directory tree to find the templates
    search_paths = [
        os.path.join(current_dir, '..', 'metadata_templates', filename),
        os.path.join(current_dir, 'metadata_templates', filename),
    ]
    
    # Also check if we're in a site-packages structure
    site_packages_paths = []
    path_parts = current_file.split(os.sep)
    for i, part in enumerate(path_parts):
        if part == 'site-packages':
            site_packages_root = os.sep.join(path_parts[:i+1])
            site_packages_paths.extend([
                os.path.join(site_packages_root, 'pysoda', 'core', 'metadata_templates', filename),
                os.path.join(site_packages_root, 'pysoda_fairdataihub_tools', 'pysoda', 'core', 'metadata_templates', filename)
            ])
    
    all_paths = search_paths + site_packages_paths
    
    for path in all_paths:
        if os.path.exists(path):
            return path
    
    # Method 4: Try to find in Electron app resources (if not using PyInstaller)
    try:
        # Look for Electron app structure
        current_path = current_dir
        while current_path and current_path != os.path.dirname(current_path):
            electron_paths = [
                os.path.join(current_path, 'resources', 'app', 'node_modules', 'pysoda', 'core', 'metadata_templates', filename),
                os.path.join(current_path, 'resources', 'pysoda', 'core', 'metadata_templates', filename),
                os.path.join(current_path, 'app', 'pysoda', 'core', 'metadata_templates', filename)
            ]
            for path in electron_paths:
                if os.path.exists(path):
                    return path
            current_path = os.path.dirname(current_path)
    except Exception:
        pass


    # Method 5: Use importlib_resources (Python 3.7+)
    try:
        from importlib import resources
        with resources.path('metadata_templates', filename) as template_path:
            if template_path.exists():
                return str(template_path)
    except (ImportError, ModuleNotFoundError):
        # Fallback to other methods if importlib_resources is not available
        pass
    

    except Exception as e:
        raise ImportError(f"Could not locate or create template file {filename}. Error: {e}")


def create_excel(soda, upload_boolean, local_destination):
    source = get_template_path("manifest.xlsx")
    destination = join(METADATA_UPLOAD_PS_PATH, SDS_FILE_MANIFEST) if upload_boolean else local_destination
    shutil.copyfile(source, destination)
    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]
    manifest = soda["dataset_metadata"]["manifest_file"]
    # validate_schema(manifest, SCHEMA_NAME_MANIFEST)
    ascii_headers = excel_columns(start_index=0)
    custom_headers_to_column = {}

    orangeFill = PatternFill(
        start_color="FFD965", end_color="FFD965", fill_type="solid"
    )

    # Load schema to get standard headers
    schema_path = os.path.join(os.path.dirname(__file__), "../../schema/manifest.json")
    with open(schema_path, "r") as f:
        schema = json_load(f)
    # The schema is an array, so get the first item's properties
    item_schema = schema["items"][0]
    standard_headers = list(item_schema["properties"].keys())

    # Write standard headers to the first row
    for idx, header in enumerate(standard_headers):
        ws1[ascii_headers[idx] + "1"] = header.replace("_", " ")
        ws1[ascii_headers[idx] + "1"].font = Font(bold=True, size=12, name="Calibri")

    row = 2
    for entry in manifest:
        # Write standard fields
        for idx, header in enumerate(standard_headers):
            value = entry.get(header, "")
            if isinstance(value, list):
                value = " ".join(value)
            ws1[ascii_headers[idx] + str(row)] = value
            ws1[ascii_headers[idx] + str(row)].font = Font(bold=False, size=11, name="Arial")

        # Handle custom fields
        for field_name, field_value in entry.items():
            if field_name in standard_headers:
                continue
            if field_name not in custom_headers_to_column:
                custom_headers_to_column[field_name] = len(custom_headers_to_column) + len(standard_headers)
                col_idx = custom_headers_to_column[field_name]
                ws1[ascii_headers[col_idx] + "1"] = field_name
                ws1[ascii_headers[col_idx] + "1"].fill = orangeFill
                ws1[ascii_headers[col_idx] + "1"].font = Font(bold=True, size=12, name="Calibri")
            col_idx = custom_headers_to_column[field_name]
            ws1[ascii_headers[col_idx] + str(row)] = field_value
            ws1[ascii_headers[col_idx] + str(row)].font = Font(bold=False, size=11, name="Arial")
        row += 1

    # Rename additional metadata header to Additional Metadata header
    # ws1[ascii_headers[len(standard_headers)] + "1"] = "Additional Metadata"

    wb.save(destination)
    size = getsize(destination)
    if upload_boolean:
        upload_metadata_file(SDS_FILE_MANIFEST, soda, destination, True)

    return {"size": size}




def load_existing_manifest_file(manifest_file_path):
   # check that a file exists at the path 
    if not os.path.exists(manifest_file_path):
        raise FileNotFoundError(f"Manifest file not found at {manifest_file_path}")
    
   # load the xlsx file and store its first row as a headers array and the rest of the rows in a data key 
    wb = load_workbook(manifest_file_path)
    ws1 = wb["Sheet1"]
    headers = []
    data = []

    for row in ws1.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)

    for row in ws1.iter_rows(min_row=2, values_only=True):
        # grab the item in row 5
        new_row = []
        for col_data in row:
            if isinstance(col_data, list):
                # space separate the list into a string
                col_data = " ".join(col_data)
            new_row.append(col_data)
        data.append(new_row)

    return {"headers": headers, "data": data}






from constants import METADATA_UPLOAD_PS_PATH, TEMPLATE_PATH
from excel_utils import rename_headers, excel_columns
from openpyxl.styles import PatternFill
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
from utils import validate_schema


# TODO: Handle optional entries when coupled with provided entries
# TODO: Handle extending columns and filling with color when more entries are provided than the template default handles
def create_excel(soda, upload, generateDestination):
    source = join(TEMPLATE_PATH, "code_description.xlsx")
    destination = join(METADATA_UPLOAD_PS_PATH, "code_description.xlsx") if upload else generateDestination
    shutil.copyfile(source, destination)

    validate_schema(soda["dataset_metadata"]["code_description"], "code_description_schema.json")


    print(destination)
    wb = load_workbook("./" +destination)
    print(wb.sheetnames)
    ws1 = wb[wb.sheetnames[0]]

    print(ws1)

    populate_input_output_information(ws1, soda)

    populate_basic_information(ws1, soda)


    populate_ten_simple_rules(ws1, soda)


    wb.save(destination)





# TODO: Handle optional entries
def populate_input_output_information(ws1, soda):
    # populate from row 27 and column 4 up to column n, depending upon the amount of items in the array for each input output information entry
    input_output_information = soda["dataset_metadata"]["input_output_information"]

    row = 27

    excel_ascii = excel_columns(start_index=3)[0]
    ws1[excel_ascii + str(row)] = input_output_information["number_of_inputs"]

    for input, column in zip(input_output_information["inputs"], excel_columns(start_index=3)):
        row = 28
        ws1[column + str(row)] = input["input_parameter_name"]
        ws1[column + str(row + 1)] = input["input parameter type"]
        ws1[column + str(row + 2)] = input["input_parameter_description"]
        ws1[column + str(row + 3)] = input["input_units"]
        ws1[column + str(row + 4)] = input["input_default_value"]

    # populate number of outputs into row 34
    row = 34
    ws1[excel_ascii + str(row)] = input_output_information["number_of_outputs"]

    # populate the outputs from row 35 - 39
    for output, column in zip(input_output_information["outputs"], excel_columns(start_index=3)):
        row = 35
        ws1[column + str(row)] = output["output_parameter_name"]
        ws1[column + str(row + 1)] = output["output_parameter_type"]
        ws1[column + str(row + 2)] = output["output_parameter_description"]
        ws1[column + str(row + 3)] = output["output_units"]
        ws1[column + str(row + 4)] = output["output_default_value"]


def populate_basic_information(ws1, soda):
    basic_information = soda["dataset_metadata"]["basic_information"]

    # fill out basic information from row 2 - 5 starting from col 3
    row = 2
    for info, column in zip(basic_information, excel_columns(start_index=3)):
        ws1[column + str(row)] = info["RRID_term"]
        ws1[column + str(row + 1)] = info["RRID_identifier"]
        ws1[column + str(row + 2)] = info["ontology_term"]
        ws1[column + str(row + 3)] = info["ontology_identifier"]


def populate_ten_simple_rules(ws1, soda):
    ten_simple_rules = soda["dataset_metadata"]["ten_simple_rules"]
    row = 8
    ascii_cols = excel_columns(start_index=3)
    for _, rule in ten_simple_rules.items():
        ws1[ascii_cols[0] + str(row)] = rule.get("Link", "")
        ws1[ascii_cols[1] + str(row)] = rule.get("Rating", "")
        ws1[ascii_cols[2] + str(row)] = rule.get("Target", "")
        ws1[ascii_cols[3] + str(row)] = rule.get("Target Justification", "")
        ws1[ascii_cols[4] + str(row)] = rule.get("Text", "")
        row += 1



soda = {
    "dataset_metadata": {
        "input_output_information": {
            "number_of_inputs": 1,
            "inputs": [
                {
                    "input_parameter_name": "ws1", 
                    "input parameter type": "worksheet",
                    "input_parameter_description": "The worksheet to write to",
                    "input_units": "N/A",
                    "input_default_value": "worksheet",
                },
                {
                    "input_parameter_name": "soda", 
                    "input parameter type": "dictionary",
                    "input_parameter_description": "The soda dictionary",
                    "input_units": "N/A",
                    "input_default_value": "soda",
                },
            ],
            "number_of_outputs": 1,
            "outputs": [
                {
                    "output_parameter_name": "None",
                    "output_parameter_type": "None",
                    "output_parameter_description": "None",
                    "output_units": "N/A",
                    "output_default_value": "N/A",
                }
            ]
        },
        "basic_information": [
            {
                "RRID_term": "RRID:SCR_016755",
                "RRID_identifier": "https://scicrunch.org/resolver/RRID:SCR_016755",
                "ontology_term": "N/A",
                "ontology_identifier": "N/A",
            },
            {
                "RRID_term": "RRID:SCR_016755",
                "RRID_identifier": "https://scicrunch.org/resolver/RRID:SCR_016755",
                "ontology_term": "N/A",
                "ontology_identifier": "N/A",
            }
        ],
        "ten_simple_rules": {
            "TSR1": {
                "Link": "https://fairsharing.org/bsg-s001134",
                "Rating": "3",
                "Target": "4",
                "Target Justification": "Silly string",
                "Text": "All the silly things with silly string"
            },
            "TSR2": {
                "Link": "https://google.com",
                "Rating": "3",
                "Target": "4",
                "Target Justification": "Silly string",
                "Text": "All the silly things with silly string"
            },
            "TSR3a": {
                "Link": "https://google.com",
                "Rating": "3",
                "Target": "4",
                "Target Justification": "Silly string",
                "Text": "All the silly things with silly string"
            },
            "TSR3b":{
                "Link": "https://google.com",
            },
            "TSR3c": {
                "Link": "https://google.com",
            },
            "TSR4": {
                "Link": "https://google.com",
                "Rating": "3",
                "Target": "4",
                "Target Justification": "Silly string",
                "Text": "All the silly things with silly string"
            },
            "TSR5":{
                "Link": "https://google.com",
                "Rating": "3",
                "Target": "4",
                "Target Justification": "Silly string",
                "Text": "All the silly things with silly string"
            },
            "TSR6": {
                "Link": "https://google.com",
                "Rating": "3",
                "Target": "4",
                "Target Justification": "Silly string",
                "Text": "All the silly things with silly string"
            },
            "TSR7a": {
                "Link": "https://google.com",
                "Rating": "3",
                "Target": "4",
                "Target Justification": "Silly string",
                "Text": "All the silly things with silly string"
            },
            "TSR7b": {
                "Link": "https://google"
            },
            "TSR7c": {
                "Link": "https://google.com",
            },
            "TSR7d": {
                "Link": "https://google.com",
            },
            "TSR7e":{
                "Link": "https://google.com",
            },
            "TSR8a": {
                "Link": "https://google.com",
                "Rating": "3",
                "Target": "4",
                "Target Justification": "Silly string",
                "Text": "All the silly things with silly string"
            },
            "TSR8b":{
                "Link": "https://google.com",
            },
            "TSR9": {
                "Link": "https://google.com",
                "Rating": "3",
                "Target": "4",
                "Target Justification": "Silly string",
                "Text": "All the silly things with silly string"
            },
            "TSR10a":{
                "Link": "https://google.com",
                "Rating": "3",
                "Target": "4",
                "Target Justification": "Silly string",
                "Text": "All the silly things with silly string"
            },
            "TSR10b": {
                "Link": "https://google.com"
            }
        }
    }
}

create_excel(soda, False, "code_description.xlsx")


    


from .constants import METADATA_UPLOAD_PS_PATH, TEMPLATE_PATH, SDS_FILE_PERFORMANCES, SCHEMA_NAME_PERFORMANCES
from .excel_utils import rename_headers, excel_columns
from openpyxl.styles import PatternFill
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
from ...utils import validate_schema
from .helpers import upload_metadata_file

def create_excel(soda, upload_boolean, local_destination):
    source = join(TEMPLATE_PATH, SDS_FILE_PERFORMANCES)

    destination = join(METADATA_UPLOAD_PS_PATH, SDS_FILE_PERFORMANCES) if upload_boolean else local_destination

    shutil.copyfile(source, destination)

    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]

    performances = soda["dataset_metadata"]["performances"]

    validate_schema(performances, SCHEMA_NAME_PERFORMANCES)

    # get the ascii column headers
    row = 2
    ascii_headers = excel_columns(start_index=0)
    for performance in performances: 
        ws1[ascii_headers[0] + str(row)] = performance.get("performance_id", "")
        ws1[ascii_headers[1] + str(row)] = performance.get("protocol_url_or_doi", "")
        ws1[ascii_headers[2] + str(row)] = performance.get("date", "")
        ws1[ascii_headers[3] + str(row)] = performance.get("start_datetime", "")
        ws1[ascii_headers[4] + str(row)] = performance.get("end_datetime", "")
        participants = " ".join(performance.get("participants", []))
        ws1[ascii_headers[5] + str(row)] = participants
        ws1[ascii_headers[6] + str(row)] = performance.get("additional_metadata", "")
        row += 1

    wb.save(destination)

    size = getsize(destination)

    ## if generating directly on Pennsieve, call upload function
    if upload_boolean:
        upload_metadata_file(SDS_FILE_PERFORMANCES, soda,  destination, True)

    return {"size": size}


# soda = {
#     "dataset_metadata": {
#         "performances": [
#             {
#                 "performance_id": "P001",
#                 "protocol_url_or_doi": "https://doi.org/10.1234/abcd",
#                 "date": "2023-10-01",
#                 "start_datetime": "2023-10-01T10:00:00Z",
#                 "end_datetime": "2023-10-01T12:00:00Z",
#                 "participants": ["sub-1", "sub-2"],
#                 "additional_metadata": "unfun metadata here"
#             },
#             {
#                 "performance_id": "P002",
#                 "protocol_url_or_doi": "https://doi.org/10.1234/abcde",
#                 "date": "2023-10-01",
#                 "start_datetime": "2023-10-01T10:00:00Z",
#                 "end_datetime": "2023-10-01T12:00:00Z",
#                 "participants": ["sub-1", "sub-2"],
#                 "additional_metadata": "fun metadata here"
#             }
#         ]
#     }
# }


# create_excel(soda, False, "performances.xlsx")
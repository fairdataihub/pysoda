from .constants import METADATA_UPLOAD_BF_PATH, TEMPLATE_PATH, SDS_FILE_RESOURCES, SCHEMA_NAME_RESOURCES
from .excel_utils import rename_headers, excel_columns
from openpyxl.styles import PatternFill
from os.path import join, getsize
from openpyxl import load_workbook
import shutil
from ...utils import validate_schema

def create_excel(soda, upload_boolean, filepath):
    source = join(TEMPLATE_PATH, SDS_FILE_RESOURCES)

    if upload_boolean:
        destination = join(METADATA_UPLOAD_BF_PATH, SDS_FILE_RESOURCES)
    else:
        destination = filepath

    shutil.copyfile(source, destination)

    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]

    resources = soda["dataset_metadata"]["resources"]

    validate_schema(resources, SCHEMA_NAME_RESOURCES)


    # get the ascii column headers
    row = 2
    ascii_headers = excel_columns(start_index=0)
    for performance in resources: 
        ws1[ascii_headers[0] + str(row)] = performance.get("rrid", "")
        ws1[ascii_headers[1] + str(row)] = performance.get("type", "")
        ws1[ascii_headers[2] + str(row)] = performance.get("name", "")
        ws1[ascii_headers[3] + str(row)] = performance.get("url", "")
        ws1[ascii_headers[4] + str(row)] = performance.get("vendor", "")
        ws1[ascii_headers[5] + str(row)] = performance.get("version", "")
        ws1[ascii_headers[6] + str(row)] = performance.get("id_in_protocol", "")
        ws1[ascii_headers[7] + str(row)] = performance.get("additional_metadata", "")

        row += 1

    wb.save(destination)



